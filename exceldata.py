import openpyxl, sys

def append_data(source,destination,fromsheet,tosheet):
    """appends data from an Source excel to a Destination excel"""
    print "Appending data from %s sheet %s to %s sheet %s" % \
          (source, fromsheet, destination, tosheet)
    
    wb_source = openpyxl.load_workbook(source)
    wb_dest = openpyxl.load_workbook(destination)

    sheet_source = wb_source.get_sheet_by_name(fromsheet)
    #minus the first row
    rows_to_copy = sheet_source.max_row
    sheet_dest = wb_dest.get_sheet_by_name(tosheet)

    source_headers = sheet_source.rows[0]
    dest_headers = sheet_dest.rows[0]

    if source_headers != dest_headers:
        print "Your files don't have the same number of columns"
        return
    
    #creates list of value to copy 
    data_to_copy = [
                    map(lambda cell: cell.value, row)
                    for row in sheet_source.rows[1:rows_to_copy]
                    ]

    #next available row in destination sheet rows start at
    current_row = sheet_dest.max_row + 1
    
    #copy data
    for row in data_to_copy:
        #move back to first column on new row
        current_column = 1
        for cell in row:
            sheet_dest.cell(row=current_row,column=current_column).value = cell
            current_column += 1
        #once a row has been added move to next row
        current_row += 1
    #save changes
    wb_dest.save(destination)
    print "Append Completed. Appended %d rows" % (rows_to_copy - 1)

def main():
    source = 'tocopy1.xlsx'
    destination = 'example.xlsx'
    fromsheet = 'Sheet1'
    tosheet = 'Sheet3'
    append_data(source, destination, fromsheet, tosheet)

if __name__ == "__main__":
    main()
